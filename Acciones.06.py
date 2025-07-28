#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de análisis y predicción de precios de acciones que integra:
  - Recolección de datos históricos, indicadores técnicos, macroeconómicos y fundamentales.
  - Análisis de sentimiento de noticias mediante FinBERT.
  - Entrenamiento de un modelo LSTM para predicción de precios (pronóstico a 1 día).
  - Procesos adicionales de validación, calibración (ajuste de hiperparámetros) y backtesting.
  - Generación de reportes analíticos y gráficos.
  - Interfaz gráfica con Tkinter.

NOTA: Este script es experimental. Se recomienda realizar un backtesting y validación robustos antes de tomar decisiones de inversión.
"""

import os, sys, time, io, logging, traceback, threading
from datetime import datetime, timedelta

import numpy as np
import pandas as pd, yfinance as yf, requests, ta
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense, Dropout, Input
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping

import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from tqdm import tqdm

import plotly.graph_objs as go
from plotly.subplots import make_subplots
import kaleido

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle

import tkinter as tk
from tkinter import scrolledtext, messagebox, ttk

# Mostrar directorio actual
print("Directorio de trabajo actual:", os.getcwd())

# Configuración de logging y supresión de advertencias de TensorFlow
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
tf.get_logger().setLevel('ERROR')

# Configuración del dispositivo
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
cuda_message = f"Usando dispositivo: {device}"
print(cuda_message)

# Parámetros globales
WINDOW_SIZE = 60
FORECAST_HORIZON = 1
MAX_EPOCHS = 200
PATIENCE = 20

# Lista de características para el modelo (14 columnas)
FEATURES = ['Close', 'Volume', 'SMA', 'EMA', 'RSI', 'RSI_7', 'MACD', 'MACD_Signal', 
            'P/E Ratio', 'P/S Ratio', 'News_Sentiment', 'Interest_Rate', 'Unemployment', 'Inflation']

# Directorios y API Keys
DESKTOP_PATH = os.path.join(os.path.expanduser("~"), "Desktop")
COTIZACIONES_PATH = os.path.join(DESKTOP_PATH, "Cotizaciones")
os.makedirs(COTIZACIONES_PATH, exist_ok=True)

API_KEY_POLYGON = "3t5QOqGIlXf60HJ0G7wQMXTyERuJbCAJ"
API_KEY_FRED = "392a61682f985a33c6cfb9d54975e580"

# Carga del modelo FinBERT
finbert_tokenizer = AutoTokenizer.from_pretrained("ProsusAI/finbert")
finbert_model = AutoModelForSequenceClassification.from_pretrained("ProsusAI/finbert").to(device)

# --- Funciones Auxiliares ---
def get_ticker_path(symbol):
    ticker_path = os.path.join(COTIZACIONES_PATH, symbol)
    os.makedirs(ticker_path, exist_ok=True)
    return ticker_path

class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget
    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)
    def flush(self):
        pass

class TqdmToGUI(tqdm):
    def __init__(self, *args, **kwargs):
        self.gui_output = kwargs.pop('gui_output', None)
        self.last_print_n = 0
        super().__init__(*args, **kwargs, file=sys.stdout)
    def display(self, msg=None, pos=None):
        if self.gui_output and self.n > self.last_print_n:
            self.last_print_n = self.n
            percentage = int(self.n / self.total * 100)
            self.gui_output.insert(tk.END, f"\rProgreso: {percentage}%\n")
            self.gui_output.see(tk.END)
            self.gui_output.update_idletasks()
    def close(self):
        self.disable = True

# --- Funciones para Análisis de Sentimiento y Recolección de Datos ---
def get_sentiment_finbert(texts, batch_size=32, gui_output=None):
    results = []
    total_batches = len(texts) // batch_size + (1 if len(texts) % batch_size != 0 else 0)
    for i in range(0, len(texts), batch_size):
        if gui_output:
            progress = (i // batch_size + 1) / total_batches
            gui_output.insert(tk.END, f"\rAnalizando sentimiento: {progress:.1%}\n")
            gui_output.update_idletasks()
        batch = texts[i:i+batch_size]
        inputs = finbert_tokenizer(batch, return_tensors="pt", padding=True, truncation=True, max_length=512)
        inputs = {k: v.to(device) for k, v in inputs.items()}
        try:
            with torch.no_grad():
                outputs = finbert_model(**inputs)
            probabilities = torch.nn.functional.softmax(outputs.logits, dim=-1)
            sentiments = torch.argmax(probabilities, dim=1)
            scores = probabilities.gather(1, sentiments.unsqueeze(1)).squeeze(1)
            for sentiment, score in zip(sentiments, scores):
                if sentiment == 0:
                    results.append(("Negativo", -score.item()))
                elif sentiment == 1:
                    results.append(("Neutral", 0))
                else:
                    results.append(("Positivo", score.item()))
        except RuntimeError as e:
            if "out of memory" in str(e):
                torch.cuda.empty_cache()
                if gui_output:
                    gui_output.insert(tk.END, "CUDA out of memory. Procesando en CPU.\n")
                inputs = {k: v.cpu() for k, v in inputs.items()}
                finbert_model.cpu()
                with torch.no_grad():
                    outputs = finbert_model(**inputs)
                finbert_model.to(device)
                probabilities = torch.nn.functional.softmax(outputs.logits, dim=-1)
                sentiments = torch.argmax(probabilities, dim=1)
                scores = probabilities.gather(1, sentiments.unsqueeze(1)).squeeze(1)
                for sentiment, score in zip(sentiments, scores):
                    if sentiment == 0:
                        results.append(("Negativo", -score.item()))
                    elif sentiment == 1:
                        results.append(("Neutral", 0))
                    else:
                        results.append(("Positivo", score.item()))
            else:
                raise e
    return results

def get_news(symbol, days=3650):
    news_file = os.path.join(get_ticker_path(symbol), f"{symbol}_news.xlsx")
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    if os.path.exists(news_file):
        existing_news = pd.read_excel(news_file)
        existing_news['date'] = pd.to_datetime(existing_news['date'])
        if not existing_news.empty:
            start_date = existing_news['date'].max() + timedelta(days=1)
        print(f"Actualizando noticias para {symbol} desde {start_date}")
    else:
        existing_news = pd.DataFrame(columns=['date', 'time', 'title', 'sentiment', 'sentiment_score'])
        print(f"Descargando todas las noticias para {symbol} desde {start_date}")
    base_url = "https://api.polygon.io/v2/reference/news"
    params = {
        "ticker": symbol,
        "order": "asc",
        "limit": 1000,
        "sort": "published_utc",
        "published_utc.gte": start_date.strftime('%Y-%m-%d'),
        "published_utc.lte": end_date.strftime('%Y-%m-%d'),
        "apiKey": API_KEY_POLYGON
    }
    news_data = []
    while True:
        try:
            response = requests.get(base_url, params=params)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"Error en la llamada a la API: {e}")
            break
        data = response.json()
        for item in data.get('results', []):
            try:
                published_date = datetime.fromisoformat(item['published_utc'].replace('Z', '+00:00'))
            except Exception as ex:
                logging.error(f"Error al parsear la fecha: {ex}")
                continue
            news_data.append({
                'date': published_date.date(),
                'time': published_date.strftime("%H:%M:%S"),
                'title': item['title']
            })
        print(f"Obtenidas {len(news_data)} noticias nuevas hasta ahora...")
        if 'next_url' in data and data['next_url']:
            params = {"apiKey": API_KEY_POLYGON}
            base_url = data['next_url']
            time.sleep(12)
        else:
            break
    new_news = pd.DataFrame(news_data)
    if not new_news.empty:
        new_news['date'] = pd.to_datetime(new_news['date'])
        new_news = new_news.sort_values('date')
    all_news = pd.concat([existing_news, new_news]).drop_duplicates().reset_index(drop=True)
    with pd.ExcelWriter(news_file, engine='openpyxl') as writer:
        all_news.to_excel(writer, index=False, sheet_name='Noticias')
    print(f"Total de noticias guardadas para {symbol}: {len(all_news)}")
    return all_news

def get_macro_indicators(fred_api_key, start_date, end_date):
    try:
        from fredapi import Fred
        fred = Fred(api_key=fred_api_key)
        interest_rate = fred.get_series('FEDFUNDS', start_date=start_date, end_date=end_date)
        cpi = fred.get_series('CPIAUCSL', start_date=start_date, end_date=end_date)
        unemployment = fred.get_series('UNRATE', start_date=start_date, end_date=end_date)
        macro_data = pd.DataFrame({
            'Interest_Rate': interest_rate,
            'CPI': cpi,
            'Unemployment': unemployment
        })
        macro_data['Inflation'] = macro_data['CPI'].pct_change(periods=12)
        macro_data = macro_data.drop('CPI', axis=1)
        macro_data.index = pd.to_datetime(macro_data.index)
        if macro_data.index.tz is not None:
            macro_data.index = macro_data.index.tz_localize(None)
        macro_data = macro_data.ffill().bfill()
        return macro_data
    except Exception as e:
        logging.error(f"Error al obtener indicadores macroeconómicos: {str(e)}")
        return None

def collect_data(symbol, start_date, end_date, fred_api_key, gui_output=None):
    """
    Recolecta y combina:
      - Datos históricos (yfinance).
      - Indicadores macroeconómicos (FRED).
      - Indicadores técnicos.
      - Datos fundamentales.
      - Noticias y análisis de sentimiento.
    
    Devuelve tres objetos:
      - train_data: DataFrame con las columnas de FEATURES (para entrenamiento, sin NaN).
      - full_data: DataFrame completo que incluye la columna 'News_Sentiment_raw' (para graficar).
      - news: DataFrame con las noticias.
    """
    try:
        if gui_output:
            gui_output.insert(tk.END, f"Recolectando datos históricos para {symbol}...\n")
        stock = yf.Ticker(symbol)
        stock_data = stock.history(start=start_date, end=end_date)
        if stock_data.empty:
            logging.error(f"No se pudieron obtener datos históricos para {symbol}")
            return None, None, None
        info = stock.info
        pe_ratio = info.get('trailingPE', np.nan)
        ps_ratio = info.get('priceToSalesTrailing12Months', np.nan)
        stock_data['P/E Ratio'] = pe_ratio
        stock_data['P/S Ratio'] = ps_ratio
        stock_data.index = pd.to_datetime(stock_data.index).tz_localize(None)
        if gui_output:
            gui_output.insert(tk.END, f"Obteniendo indicadores macroeconómicos para {symbol}...\n")
        macro_data = get_macro_indicators(fred_api_key, start_date, end_date)
        if macro_data is not None:
            macro_data.index = pd.to_datetime(macro_data.index).tz_localize(None)
        data = stock_data[['Close', 'Volume', 'P/E Ratio', 'P/S Ratio']].copy()
        if macro_data is not None:
            data = data.join(macro_data, how='left')
        if gui_output:
            gui_output.insert(tk.END, f"Calculando indicadores técnicos para {symbol}...\n")
        data['SMA'] = ta.trend.sma_indicator(data['Close'], window=20)
        data['EMA'] = ta.trend.ema_indicator(data['Close'], window=20)
        data['RSI'] = ta.momentum.rsi(data['Close'], window=14)
        data['RSI_7'] = ta.momentum.rsi(data['Close'], window=7)
        macd = ta.trend.MACD(data['Close'])
        data['MACD'] = macd.macd()
        data['MACD_Signal'] = macd.macd_signal()
        if gui_output:
            gui_output.insert(tk.END, f"Recolectando noticias y analizando sentimiento para {symbol}...\n")
        news = get_news(symbol, days=(end_date - start_date).days)
        daily_sentiment = None
        if not news.empty:
            sentiments = get_sentiment_finbert(news['title'].tolist(), gui_output=gui_output)
            news['sentiment'], news['sentiment_score'] = zip(*sentiments)
            daily_sentiment = news.groupby('date')['sentiment_score'].mean()
        # Preparar la columna de sentimiento
        sentiment_data = pd.DataFrame(index=data.index)
        if daily_sentiment is not None and not daily_sentiment.empty:
            # Para entrenamiento, rellenamos con 0 los días sin dato
            sentiment_data['News_Sentiment'] = daily_sentiment.fillna(0)
        else:
            if gui_output:
                gui_output.insert(tk.END, f"No se encontraron noticias para {symbol}. Se usará 0 para 'News_Sentiment'.\n")
            sentiment_data['News_Sentiment'] = 0
        # Conservar la versión "raw" sin rellenar para graficar
        raw_sentiment = daily_sentiment  # Puede contener NaN
        # Aplicar fill fwd y bwd al resto de columnas
        data = data.ffill().bfill()
        # Unir la columna de sentimiento para entrenamiento (rellenada) y la raw
        data['News_Sentiment'] = sentiment_data['News_Sentiment']
        data['News_Sentiment_raw'] = raw_sentiment  # se mantienen NaN en días sin dato
        # Finalmente, asegurarse de que el conjunto de entrenamiento no tenga NaN (solo en las columnas de FEATURES)
        train_data = data[FEATURES].fillna(0)
        full_data = data.copy()
        if gui_output:
            gui_output.insert(tk.END, f"Datos procesados para {symbol}. Total filas: {len(train_data)}\n")
        return train_data, full_data, news
    except Exception as e:
        logging.error(f"Error al recolectar datos para {symbol}: {str(e)}")
        logging.error(traceback.format_exc())
        if gui_output:
            gui_output.insert(tk.END, f"Error en la recolección para {symbol}: {str(e)}\n")
        return None, None, None

# --- Funciones para Preparación, Entrenamiento, Calibración y Backtesting ---
def calculate_sample_weights(X):
    num_samples = X.shape[0]
    weights = np.exp(np.linspace(0, 1, num_samples)) - 1
    return weights / np.mean(weights)

def preprocess_data(data, window_size=WINDOW_SIZE, forecast_horizon=FORECAST_HORIZON):
    scaler = MinMaxScaler()
    scaled_data = scaler.fit_transform(data)
    n_samples = len(scaled_data) - window_size - forecast_horizon + 1
    if n_samples <= 0:
        raise ValueError("No hay suficientes datos para la ventana y horizonte especificados.")
    X, y = [], []
    for i in range(n_samples):
        X.append(scaled_data[i:i+window_size, :])
        y.append(scaled_data[i+window_size:i+window_size+forecast_horizon, 0])
    X, y = np.array(X), np.array(y)
    sample_weights = calculate_sample_weights(X)
    return X, y, scaler, sample_weights

def create_model(input_shape, lstm_units=150, dropout_rate=0.3, learning_rate=0.0005, forecast_horizon=FORECAST_HORIZON):
    model = Sequential([
        Input(shape=input_shape),
        LSTM(lstm_units, return_sequences=True),
        Dropout(dropout_rate),
        LSTM(lstm_units, return_sequences=False),
        Dropout(dropout_rate),
        Dense(forecast_horizon)
    ])
    model.compile(optimizer=Adam(learning_rate=learning_rate), loss='mse')
    return model

def generate_predictions(model, latest_data, scaler, forecast_horizon=FORECAST_HORIZON, window_size=WINDOW_SIZE, gui_output=None):
    if gui_output:
        gui_output.insert(tk.END, f"Generando predicción directa para los próximos {forecast_horizon} día(s)...\n")
        gui_output.update_idletasks()
    scaled_latest = scaler.transform(latest_data)
    X_input = scaled_latest[-window_size:].reshape(1, window_size, scaled_latest.shape[1])
    pred = model.predict(X_input, verbose=0)[0]
    dummy = np.zeros((forecast_horizon, scaler.scale_.shape[0]))
    dummy[:, 0] = pred
    unscaled_pred = scaler.inverse_transform(dummy)[:, 0]
    if gui_output:
        gui_output.insert(tk.END, "Predicción completada.\n")
        gui_output.update_idletasks()
    return unscaled_pred

class ProgressCallback(tf.keras.callbacks.Callback):
    def __init__(self, num_epochs, gui_output, progress_bar, early_stopping_text):
        self.num_epochs = num_epochs
        self.gui_output = gui_output
        self.progress_bar = progress_bar
        self.early_stopping_text = early_stopping_text
        self.last_update = 0
        self.best_val_loss = float('inf')
        self.worst_val_loss = float('-inf')
    def on_epoch_end(self, epoch, logs=None):
        if self.gui_output and epoch % 5 == 0:
            self.gui_output.insert(tk.END, f"Época {epoch + 1} completada.\n")
            self.gui_output.see(tk.END)
        if self.progress_bar:
            progress = int(((epoch + 1) / self.num_epochs) * 100)
            if progress > self.last_update:
                self.progress_bar['value'] = progress
                self.progress_bar.update()
                self.last_update = progress
        if logs and 'val_loss' in logs:
            current_val_loss = logs['val_loss']
            self.best_val_loss = min(self.best_val_loss, current_val_loss)
            self.worst_val_loss = max(self.worst_val_loss, current_val_loss)
            self.early_stopping_text.set(f"Early Stopping. Mejor: {self.best_val_loss:.6f} | Actual: {current_val_loss:.6f} | Peor: {self.worst_val_loss:.6f}")

def train_model(X, y, sample_weights, gui_output=None, progress_bar=None, early_stopping_text=None, epochs=MAX_EPOCHS):
    split_index = int(len(X) * 0.8)
    X_train, X_test = X[:split_index], X[split_index:]
    y_train, y_test = y[:split_index], y[split_index:]
    weights_train = sample_weights[:split_index]
    model = create_model((X_train.shape[1], X_train.shape[2]), forecast_horizon=FORECAST_HORIZON)
    early_stopping = EarlyStopping(monitor='val_loss', patience=PATIENCE, restore_best_weights=True)
    progress_callback = ProgressCallback(epochs, gui_output, progress_bar, early_stopping_text)
    history = model.fit(X_train, y_train, sample_weight=weights_train, epochs=epochs, batch_size=32,
                        validation_split=0.1, callbacks=[early_stopping, progress_callback], verbose=0)
    return model, X_test, y_test, history

def calibrate_model(X, y, sample_weights, param_grid, epochs=30, validation_split=0.1, batch_size=32, gui_output=None):
    best_val_loss = float('inf')
    best_params = None
    best_model = None
    best_history = None
    split_index = int(len(X) * (1 - validation_split))
    X_train, X_val = X[:split_index], X[split_index:]
    y_train, y_val = y[:split_index], y[split_index:]
    weights_train = sample_weights[:split_index]
    for units in param_grid.get('lstm_units', [150]):
        for dropout in param_grid.get('dropout_rate', [0.3]):
            for lr in param_grid.get('learning_rate', [0.0005]):
                if gui_output:
                    gui_output.insert(tk.END, f"Probando: LSTM_units={units}, dropout={dropout}, learning_rate={lr}\n")
                    gui_output.update_idletasks()
                model = create_model((X_train.shape[1], X_train.shape[2]), lstm_units=units, dropout_rate=dropout, learning_rate=lr, forecast_horizon=FORECAST_HORIZON)
                history = model.fit(X_train, y_train, sample_weight=weights_train, epochs=epochs, batch_size=batch_size, validation_data=(X_val, y_val), verbose=0)
                final_val_loss = history.history['val_loss'][-1]
                if gui_output:
                    gui_output.insert(tk.END, f"Resultado: val_loss={final_val_loss}\n")
                    gui_output.update_idletasks()
                if final_val_loss < best_val_loss:
                    best_val_loss = final_val_loss
                    best_params = {'lstm_units': units, 'dropout_rate': dropout, 'learning_rate': lr}
                    best_model = model
                    best_history = history
    if gui_output:
        gui_output.insert(tk.END, f"Mejores parámetros: {best_params} con val_loss={best_val_loss}\n")
        gui_output.update_idletasks()
    return best_model, best_params, best_history

def backtest_model(data, scaler, window_size, model, step=1):
    predictions, actuals, indices = [], [], []
    for i in range(window_size, len(data), step):
        window_data = data.iloc[i-window_size:i]
        scaled_window = scaler.transform(window_data)
        X_input = scaled_window.reshape(1, window_size, scaled_window.shape[1])
        pred = model.predict(X_input, verbose=0)
        dummy = np.zeros((1, scaler.scale_.shape[0]))
        dummy[0, 0] = pred[0, 0]
        unscaled_pred = scaler.inverse_transform(dummy)[0, 0]
        predictions.append(unscaled_pred)
        actuals.append(data['Close'].iloc[i])
        indices.append(data.index[i])
    rmse = np.sqrt(mean_squared_error(actuals, predictions))
    mae = mean_absolute_error(actuals, predictions)
    r2 = r2_score(actuals, predictions)
    return indices, predictions, actuals, rmse, mae, r2

def plot_backtest_results(indices, predictions, actuals, symbol):
    fig = make_subplots(rows=1, cols=1)
    fig.add_trace(go.Scatter(x=indices, y=actuals, mode='lines', name='Actual', line=dict(color='green')))
    fig.add_trace(go.Scatter(x=indices, y=predictions, mode='lines', name='Predicción', line=dict(color='red', dash='dash')))
    fig.update_layout(title=f'Backtesting de predicciones para {symbol}', xaxis_title='Fecha', yaxis_title='Precio', width=1600, height=800)
    return fig

def generate_analytical_report(symbol, model, history, rmse, mae, r2, predictions, current_price, feature_importance, sensitivity, news):
    report = f"""
Análisis Predictivo para {symbol}

1. Entrenamiento del modelo:
   - Épocas entrenadas: {len(history.epoch)}
   - Pérdida de validación final: {history.history['val_loss'][-1]:.6f}
   - {'Early stopping activado' if len(history.epoch) < MAX_EPOCHS else 'Entrenamiento completo'}

2. Evaluación del modelo:
   - RMSE: {rmse:.4f}
   - MAE: {mae:.4f}
   - R2 Score: {r2:.4f}
   (El R2 indica que se explica aproximadamente el {r2*100:.2f}% de la variabilidad.)

3. Predicción (forecast multi-step, 1 día):
   - Precio actual: ${current_price:.2f}
   - Predicción para el próximo día: ${predictions[0]:.2f}
    """
    report += "\n4. Análisis de características:\n   a. Importancia:\n"
    for feat, imp in sorted(zip(FEATURES, feature_importance), key=lambda x: x[1], reverse=True):
        report += f"      - {feat}: {imp:.4f}\n"
    report += "\n   b. Sensibilidad:\n"
    total_sens = sum(abs(s) for s in sensitivity)
    for feat, sens in sorted(zip(FEATURES, sensitivity), key=lambda x: abs(x[1]), reverse=True):
        perc = (abs(sens)/total_sens)*100 if total_sens else 0
        report += f"      - {feat}: {sens:.4f} ({perc:.2f}%)\n"
    report += "\n5. Noticias recientes:\n"
    if news is not None and not news.empty:
        for _, art in news.head().iterrows():
            report += f"      - {art['title']} (Sentimiento: {art['sentiment']}, Score: {art['sentiment_score']:.2f})\n"
    else:
        report += "      No se pudieron obtener noticias recientes.\n"
    report += f"""
6. Conclusiones:
   - El modelo presenta un rendimiento {'excelente' if r2 > 0.9 else 'bueno' if r2 > 0.8 else 'aceptable' if r2 > 0.7 else 'pobre'}.
   - Las variables fundamentales (P/E y P/S) se integran junto con indicadores técnicos y macroeconómicos.
   - Se recomienda usar este análisis como herramienta complementaria, realizando además un backtesting extenso y ajustes adicionales.
    
7. Recomendaciones:
   - Complementar el análisis con información fundamental adicional.
   - Reentrenar periódicamente con datos actualizados.
   - Validar la estrategia en un entorno de backtesting antes de tomar decisiones de inversión.
    """
    return report

def explain_model(model, X, feature_names, gui_output=None):
    if gui_output:
        gui_output.insert(tk.END, "Calculando importancia mediante gradientes...\n")
        gui_output.update_idletasks()
    try:
        x_tensor = tf.convert_to_tensor(X[0:1], dtype=tf.float32)
        with tf.GradientTape() as tape:
            tape.watch(x_tensor)
            y_pred = model(x_tensor, training=False)
        gradients = tape.gradient(y_pred, x_tensor)
        avg_grad = tf.reduce_mean(tf.abs(gradients), axis=1).numpy()[0]
        if np.sum(avg_grad) != 0:
            feature_importance = avg_grad / np.sum(avg_grad)
        else:
            feature_importance = np.ones_like(avg_grad) / len(avg_grad)
    except Exception as e:
        if gui_output:
            gui_output.insert(tk.END, f"Error en gradientes: {str(e)}\n")
            gui_output.update_idletasks()
        feature_importance = np.ones(X.shape[2]) / X.shape[2]
    sensitivity = []
    base_pred = model.predict(X[0:1], verbose=0)
    for i in range(X.shape[2]):
        X_mod = X[0:1].copy()
        X_mod[0, :, i] *= 1.1
        new_pred = model.predict(X_mod, verbose=0)
        sensitivity.append((new_pred - base_pred)[0][0])
        if gui_output and (i+1) % 5 == 0:
            gui_output.insert(tk.END, f"Análisis de sensibilidad: {i+1}/{X.shape[2]} completado.\n")
            gui_output.update_idletasks()
    if gui_output:
        gui_output.insert(tk.END, "Análisis de características completado.\n")
        gui_output.update_idletasks()
    return feature_importance, sensitivity

def generate_comparison_report(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación de Acciones"
    headers = ["Símbolo", "Último Precio", "Predicción (1 día)", "RMSE", "MAE", "R2 Score"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    row = 2
    for symbol, data in results.items():
        if 'report' not in data:
            continue
        rep_lines = data['report'].split('\n')
        try:
            last_price = float(next(line for line in rep_lines if "Precio actual:" in line).split('$')[1].split()[0])
            pred_val = float(next(line for line in rep_lines if "Predicción" in line).split("$")[1].split()[0])
            rmse = float(next(line for line in rep_lines if "RMSE:" in line).split(':')[1].strip())
            mae = float(next(line for line in rep_lines if "MAE:" in line).split(':')[1].strip())
            r2 = float(next(line for line in rep_lines if "R2 Score:" in line).split(':')[1].strip())
        except Exception as e:
            logging.error(f"Error procesando informe para {symbol}: {str(e)}")
            continue
        ws.cell(row=row, column=1, value=symbol)
        ws.cell(row=row, column=2, value=last_price)
        ws.cell(row=row, column=3, value=f"{pred_val:.2f}")
        ws.cell(row=row, column=4, value=rmse)
        ws.cell(row=row, column=5, value=mae)
        ws.cell(row=row, column=6, value=r2)
        row += 1
    fig = go.Figure()
    for symbol, data in results.items():
        if 'data' in data:
            fig.add_trace(go.Scatter(x=data['data'].index, y=data['data']['Close'], mode='lines', name=symbol))
    fig.update_layout(title='Comparación de Precios de Acciones', xaxis_title='Fecha', yaxis_title='Precio')
    img_buf = io.BytesIO()
    fig.write_image(img_buf, format="png")
    img = Image(img_buf)
    ws.add_image(img, 'H2')
    explanations = ["RMSE: Raíz del Error Cuadrático Medio.", "MAE: Error Absoluto Medio.", "R2 Score: Proporción de varianza explicada."]
    for i, line in enumerate(explanations, start=row+2):
        ws.cell(row=i, column=1, value=line)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    symbols_joined = "_".join(sorted(results.keys()))
    first_symbol = next(iter(results.keys()))
    ticker_path = get_ticker_path(first_symbol)
    excel_file = os.path.join(ticker_path, f"stock_comparison_{symbols_joined}_{timestamp}.xlsx")
    wb.save(excel_file)
    return excel_file

def generate_excel_report(symbol, historical_data, model_results, macro_data, news_data, pred_fig, imp_fig, price_fig, sent_price_fig, training_history):
    ticker_path = get_ticker_path(symbol)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    excel_file = os.path.join(ticker_path, f"{symbol}_analysis_{timestamp}.xlsx")
    wb = Workbook()
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'DD/MM/YYYY'
    ws1 = wb.active
    ws1.title = "Datos Históricos"
    ws1.cell(row=1, column=1, value="Fecha")
    for col, header in enumerate(historical_data.columns, start=2):
        ws1.cell(row=1, column=col, value=header)
    for row, (index, values) in enumerate(historical_data.iterrows(), start=2):
        cell = ws1.cell(row=row, column=1, value=index)
        if isinstance(index, (datetime, pd.Timestamp)):
            cell.style = date_style
        for col, value in enumerate(values, start=2):
            ws1.cell(row=row, column=col, value=value)
    ws2 = wb.create_sheet(title="Informe Analítico")
    ws2.column_dimensions['A'].width = 100
    for row, line in enumerate(model_results.split('\n'), start=1):
        ws2.cell(row=row, column=1, value=line)
    ws3 = wb.create_sheet(title="Datos Macroeconómicos")
    ws3.cell(row=1, column=1, value="Fecha")
    for col, header in enumerate(macro_data.columns, start=2):
        ws3.cell(row=1, column=col, value=header)
    for row, (index, values) in enumerate(macro_data.iterrows(), start=2):
        cell = ws3.cell(row=row, column=1, value=index)
        if isinstance(index, (datetime, pd.Timestamp)):
            cell.style = date_style
        for col, value in enumerate(values, start=2):
            ws3.cell(row=row, column=col, value=value)
    if news_data is not None and not news_data.empty:
        ws4 = wb.create_sheet(title="Noticias")
        headers = ["Fecha", "Hora", "Título", "Sentimiento", "Puntuación de Sentimiento"]
        for col, header in enumerate(headers, start=1):
            ws4.cell(row=1, column=col, value=header)
        for row, (_, values) in enumerate(news_data.iterrows(), start=2):
            date_value = values.get('date', '')
            if isinstance(date_value, (datetime, pd.Timestamp)):
                cell = ws4.cell(row=row, column=1, value=date_value)
                cell.style = date_style
            else:
                ws4.cell(row=row, column=1, value=date_value)
            ws4.cell(row=row, column=2, value=values.get('time', ''))
            ws4.cell(row=row, column=3, value=values.get('title', ''))
            ws4.cell(row=row, column=4, value=values.get('sentiment', ''))
            ws4.cell(row=row, column=5, value=values.get('sentiment_score', ''))
    if pred_fig is not None:
        img_buf = io.BytesIO()
        pred_fig.write_image(img_buf, format="png")
        img = Image(img_buf)
        ws5 = wb.create_sheet(title="Gráfico de Predicciones")
        ws5.add_image(img, 'A1')
    if imp_fig is not None:
        img_buf = io.BytesIO()
        imp_fig.write_image(img_buf, format="png")
        img = Image(img_buf)
        ws6 = wb.create_sheet(title="Importancia de Características")
        ws6.add_image(img, 'A1')
        feature_definitions = {
            'Close': 'Precio de cierre diario.',
            'Volume': 'Número de acciones negociadas.',
            'SMA': 'Media Móvil Simple.',
            'EMA': 'Media Móvil Exponencial.',
            'RSI': 'Índice de Fuerza Relativa (14 días).',
            'RSI_7': 'Índice de Fuerza Relativa (7 días).',
            'MACD': 'Convergencia/Divergencia de Medias Móviles.',
            'MACD_Signal': 'Señal del MACD.',
            'P/E Ratio': 'Ratio Precio/Beneficio.',
            'P/S Ratio': 'Ratio Precio/Ventas.',
            'News_Sentiment': 'Sentimiento de noticias (relleno).',
            'Interest_Rate': 'Tasa de interés de referencia.',
            'Unemployment': 'Tasa de desempleo.',
            'Inflation': 'Tasa de inflación anual.'
        }
        for row, (feat, defin) in enumerate(feature_definitions.items(), start=1):
            ws6.cell(row=row, column=12, value=feat)
            ws6.cell(row=row, column=13, value=defin)
        explanation = [
            "La importancia se calcula mediante el promedio absoluto de los gradientes del output respecto a la entrada.",
            "La sensibilidad se calcula perturbando cada característica.",
            "Estas métricas son aproximaciones para interpretar el modelo LSTM."
        ]
        for i, line in enumerate(explanation):
            ws6.cell(row=28+i, column=1, value=line)
    if price_fig is not None:
        img_buf = io.BytesIO()
        price_fig.write_image(img_buf, format="png")
        img = Image(img_buf)
        ws7 = wb.create_sheet(title="Gráfico de Precios")
        ws7.add_image(img, 'A1')
    if sent_price_fig is not None:
        img_buf = io.BytesIO()
        sent_price_fig.write_image(img_buf, format="png")
        img = Image(img_buf)
        ws8 = wb.create_sheet(title="Sentimiento vs Precio")
        ws8.add_image(img, 'A1')
    if training_history is not None and 'loss' in training_history.history and 'val_loss' in training_history.history:
        ws9 = wb.create_sheet(title="Entrenamiento del Modelo")
        train_loss = training_history.history['loss']
        val_loss = training_history.history['val_loss']
        epochs_list = list(range(1, len(train_loss)+1))
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=epochs_list, y=train_loss, mode='lines', name='Pérdida de entrenamiento'))
        fig.add_trace(go.Scatter(x=epochs_list, y=val_loss, mode='lines', name='Pérdida de validación'))
        fig.update_layout(title='Entrenamiento del Modelo', xaxis_title='Época', yaxis_title='Pérdida', width=800, height=600)
        img_buf = io.BytesIO()
        fig.write_image(img_buf, format="png")
        img = Image(img_buf)
        ws9.add_image(img, 'A1')
        explanation = [
            "La línea azul indica la pérdida en el entrenamiento y la naranja en validación.",
            "El Early Stopping se activa cuando la pérdida de validación no mejora.",
            "Es recomendable evaluar el modelo en datos no vistos."
        ]
        for i, line in enumerate(explanation):
            ws9.cell(row=28+i, column=1, value=line)
    else:
        print("No se pudo generar el gráfico de entrenamiento.")
    wb.save(excel_file)
    return excel_file

def plot_predictions(symbol, data, predictions):
    fig = make_subplots(rows=1, cols=1)
    fig.add_trace(go.Scatter(x=data.index[-30:], y=data['Close'][-30:], mode='lines', name='Datos Históricos', line=dict(color='blue')))
    last_date = data.index[-1]
    last_price = data['Close'].iloc[-1]
    future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=len(predictions))
    pred_dates = [last_date] + list(future_dates)
    pred_values = [last_price] + list(predictions)
    fig.add_trace(go.Scatter(x=pred_dates, y=pred_values, mode='lines+markers', name='Predicciones', line=dict(color='red', dash='dash')))
    fig.update_layout(title=f'Predicciones de precios para {symbol}', xaxis_title='Fecha', yaxis_title='Precio', hovermode='x unified', width=1600, height=800)
    return fig

def plot_feature_importance(feature_names, importances):
    fig = go.Figure(go.Bar(x=importances, y=feature_names, orientation='h'))
    fig.update_layout(title='Importancia de las características', xaxis_title='Importancia', yaxis_title='Características', yaxis=dict(autorange="reversed"))
    return fig

def plot_sentiment_vs_price(sentiment_series, price_series, symbol):
    sentiment_series.index = pd.to_datetime(sentiment_series.index)
    price_series.index = pd.to_datetime(price_series.index)
    valid_sentiment = sentiment_series.dropna()
    if valid_sentiment.empty:
        print(f"No hay datos de sentimiento válidos para {symbol}")
        return None, None
    common_dates = valid_sentiment.index.intersection(price_series.index)
    if common_dates.empty:
        print(f"No hay fechas comunes entre sentimiento y precio para {symbol}")
        return None, None
    df = pd.DataFrame({'Sentiment': valid_sentiment.loc[common_dates], 'Price': price_series.loc[common_dates]}).dropna()
    x_full = price_series.index.map(datetime.toordinal).values
    z_full = np.polyfit(x_full, price_series.values, 1)
    p_full = np.poly1d(z_full)
    trend_full = p_full(x_full)
    fig1 = make_subplots(specs=[[{"secondary_y": False}]])
    fig1.add_trace(go.Scatter(x=price_series.index, y=price_series, name="Precio", mode='lines', line=dict(color='green')), secondary_y=False)
    fig1.add_trace(go.Scatter(x=price_series.index, y=trend_full, name="Tendencia Precio", line=dict(color='red', dash='dash')), secondary_y=False)
    fig1.update_layout(title_text=f"Precio para {symbol} (Período completo)", xaxis_title="Fecha", yaxis_title="Precio", width=1800, height=600)
    x_dates = df.index.map(datetime.toordinal).values
    z_sent = np.polyfit(x_dates, df['Sentiment'].astype(float).values, 1)
    p_sent = np.poly1d(z_sent)
    trend_sent = p_sent(x_dates)
    z_price = np.polyfit(x_dates, df['Price'].astype(float).values, 1)
    p_price = np.poly1d(z_price)
    trend_price = p_price(x_dates)
    fig2 = make_subplots(specs=[[{"secondary_y": True}]])
    fig2.add_trace(go.Scatter(x=df.index, y=df['Sentiment'], name="Sentimiento", mode='lines', line=dict(color='blue')), secondary_y=False)
    fig2.add_trace(go.Scatter(x=df.index, y=df['Price'], name="Precio", mode='lines', line=dict(color='green')), secondary_y=True)
    if len(df) > 1:
        fig2.add_trace(go.Scatter(x=df.index, y=trend_sent, name="Tendencia Sentimiento", line=dict(color='blue', dash='dash')), secondary_y=False)
        fig2.add_trace(go.Scatter(x=df.index, y=trend_price, name="Tendencia Precio", line=dict(color='red', dash='dash')), secondary_y=True)
    fig2.update_layout(title_text=f"Sentimiento vs Precio para {symbol} (días con dato real)", xaxis_title="Fecha", width=1800, height=600)
    fig2.update_yaxes(title_text="Sentimiento", secondary_y=False)
    fig2.update_yaxes(title_text="Precio", secondary_y=True)
    return fig1, fig2

def evaluate_model(model, X_test, y_test, scaler, gui_output=None):
    if gui_output:
        gui_output.insert(tk.END, "Evaluando el modelo...\n")
        gui_output.update_idletasks()
    y_pred = model.predict(X_test, verbose=0)
    n_features = scaler.scale_.shape[0]
    y_pred_rescaled, y_test_rescaled = [], []
    for pred_vec, true_vec in zip(y_pred, y_test):
        dummy_pred = np.zeros((FORECAST_HORIZON, n_features))
        dummy_true = np.zeros((FORECAST_HORIZON, n_features))
        dummy_pred[:, 0] = pred_vec
        dummy_true[:, 0] = true_vec
        y_pred_rescaled.append(scaler.inverse_transform(dummy_pred)[:, 0])
        y_test_rescaled.append(scaler.inverse_transform(dummy_true)[:, 0])
    y_pred_rescaled = np.array(y_pred_rescaled)
    y_test_rescaled = np.array(y_test_rescaled)
    rmse = np.sqrt(mean_squared_error(y_test_rescaled, y_pred_rescaled))
    mae = mean_absolute_error(y_test_rescaled, y_pred_rescaled)
    r2 = r2_score(y_test_rescaled, y_pred_rescaled)
    if gui_output:
        gui_output.insert(tk.END, "Evaluación completada.\n")
        gui_output.insert(tk.END, f"RMSE: {rmse:.4f}\nMAE: {mae:.4f}\nR2 Score: {r2:.4f}\n")
        gui_output.update_idletasks()
    return rmse, mae, r2, y_test_rescaled, y_pred_rescaled

def analyze_stocks(symbols, gui_output, progress_bar, early_stopping_text):
    results = {}
    for symbol in symbols:
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=3650)
            gui_output.insert(tk.END, f"Iniciando análisis para {symbol}...\n")
            train_data, full_data, news = collect_data(symbol, start_date, end_date, API_KEY_FRED, gui_output)
            if train_data is None or train_data.empty:
                gui_output.insert(tk.END, f"No se pudieron obtener datos para {symbol}.\n")
                continue
            gui_output.insert(tk.END, f"Preprocesando datos para {symbol}...\n")
            X, y, scaler, sample_weights = preprocess_data(train_data, window_size=WINDOW_SIZE, forecast_horizon=FORECAST_HORIZON)
            gui_output.insert(tk.END, f"Entrenando modelo para {symbol}...\n")
            model, X_test, y_test, history = train_model(X, y, sample_weights, gui_output, progress_bar, early_stopping_text)
            gui_output.insert(tk.END, f"Evaluando el modelo para {symbol}...\n")
            try:
                rmse, mae, r2, y_test_rescaled, y_pred_rescaled = evaluate_model(model, X_test, y_test, scaler, gui_output)
            except ValueError as ve:
                gui_output.insert(tk.END, f"Error en evaluación para {symbol}: {str(ve)}\n")
                continue
            gui_output.insert(tk.END, f"Generando predicción multi-step para {symbol}...\n")
            latest_data = train_data.iloc[-WINDOW_SIZE:, :]
            predictions = generate_predictions(model, latest_data, scaler, forecast_horizon=FORECAST_HORIZON, window_size=WINDOW_SIZE, gui_output=gui_output)
            gui_output.insert(tk.END, f"Analizando importancia de características para {symbol}...\n")
            feature_importance, sensitivity = explain_model(model, X, train_data.columns.tolist(), gui_output)
            gui_output.insert(tk.END, f"Generando informe analítico para {symbol}...\n")
            report = generate_analytical_report(symbol, model, history, rmse, mae, r2, predictions, train_data['Close'].iloc[-1], feature_importance, sensitivity, news)
            gui_output.insert(tk.END, f"Generando gráficos para {symbol}...\n")
            try:
                pred_fig = plot_predictions(symbol, train_data, predictions)
            except Exception as e:
                gui_output.insert(tk.END, f"Error en gráfico de predicciones para {symbol}: {str(e)}. Se omite.\n")
                pred_fig = None
            try:
                imp_fig = plot_feature_importance(FEATURES, feature_importance)
            except Exception as e:
                gui_output.insert(tk.END, f"Error en gráfico de importancia para {symbol}: {str(e)}. Se omite.\n")
                imp_fig = None
            try:
                sentiment_series = full_data['News_Sentiment_raw']
                price_fig, sent_price_fig = plot_sentiment_vs_price(sentiment_series, full_data['Close'], symbol)
            except Exception as e:
                gui_output.insert(tk.END, f"Error en gráfico de sentimiento vs precio para {symbol}: {str(e)}. Se omite.\n")
                price_fig, sent_price_fig = None, None
            results[symbol] = {
                'data': full_data,
                'report': report,
                'predictions': predictions,
                'figures': {
                    'predictions': pred_fig,
                    'importance': imp_fig,
                    'price': price_fig,
                    'sentiment_price': sent_price_fig
                },
                'training_history': history
            }
            try:
                excel_file = generate_excel_report(symbol, full_data, report, full_data[['Interest_Rate', 'Inflation', 'Unemployment']], news, 
                                                   pred_fig, imp_fig, price_fig, sent_price_fig, history)
                gui_output.insert(tk.END, f"Análisis para {symbol} completado. Resultados guardados en: {excel_file}\n")
            except Exception as e:
                gui_output.insert(tk.END, f"Error al generar Excel para {symbol}: {str(e)}\nDetalles: {traceback.format_exc()}\n")
        except Exception as e:
            gui_output.insert(tk.END, f"Error durante el análisis de {symbol}: {str(e)}\n")
            logging.error(f"Error en análisis de {symbol}: {str(e)}")
            logging.error(traceback.format_exc())
    return results

# --- Interfaz Gráfica con Tkinter ---
class StockPredictionGUI:
    def __init__(self, master):
        self.master = master
        master.title("Predictor de Precios de Acciones 2024. LBT 1.0")
        master.geometry("800x600")
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
        csv_path = os.path.join(application_path, "data", "SP500.csv")
        self.dir_label = tk.Label(master, text=f"Directorio de trabajo: {application_path}")
        self.dir_label.pack()
        try:
            encodings = ['utf-8', 'iso-8859-1', 'windows-1252']
            for encoding in encodings:
                try:
                    self.sp500 = pd.read_csv(csv_path, sep=";", encoding=encoding)
                    print(f"SP500.csv cargado con {encoding}")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("No se pudo decodificar SP500.csv.")
        except Exception as e:
            print(f"Error al cargar SP500.csv: {str(e)}")
            self.sp500 = pd.DataFrame(columns=['Ticker', 'Empresa'])
        self.label = tk.Label(master, text="Seleccione uno o más tickers o ingréselos separados por comas:")
        self.label.pack()
        list_frame = tk.Frame(master)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE, width=50, height=10)
        if not self.sp500.empty:
            for _, row in self.sp500.iterrows():
                self.listbox.insert(tk.END, f"{row['Ticker']} - {row['Empresa']}")
        else:
            self.listbox.insert(tk.END, "No se cargó la lista S&P 500")
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical")
        scrollbar.config(command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)
        button_frame = tk.Frame(master)
        button_frame.pack(fill=tk.X, padx=10, pady=5)
        self.predict_button = tk.Button(button_frame, text="Predecir", command=self.predict)
        self.predict_button.pack(side=tk.LEFT, expand=True)
        self.calibrate_button = tk.Button(button_frame, text="Calibrar", command=self.calibrate)
        self.calibrate_button.pack(side=tk.LEFT, expand=True)
        self.backtest_button = tk.Button(button_frame, text="Backtest", command=self.backtest)
        self.backtest_button.pack(side=tk.LEFT, expand=True)
        self.readme_button = tk.Button(button_frame, text="Léame", command=self.show_readme)
        self.readme_button.pack(side=tk.RIGHT, expand=True)
        self.entry = tk.Entry(master, width=50)
        self.entry.pack(pady=5)
        self.entry.bind('<Return>', self.predict)
        self.progress_frame = tk.Frame(master)
        self.progress_frame.pack(fill=tk.X, padx=10, pady=5)
        self.progress = ttk.Progressbar(self.progress_frame, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(expand=True)
        self.early_stopping_frame = tk.Frame(master)
        self.early_stopping_frame.pack(fill=tk.X, padx=10, pady=5)
        self.early_stopping_text = tk.StringVar()
        self.early_stopping_text.set("Early Stopping. Mejor: N/A | Actual: N/A | Peor: N/A")
        self.early_stopping_label = tk.Label(self.early_stopping_frame, textvariable=self.early_stopping_text)
        self.early_stopping_label.pack()
        self.progress_label = tk.Label(master, text="La barra muestra el progreso (máximo 200 épocas).", wraplength=780)
        self.progress_label.pack(pady=5)
        self.output = scrolledtext.ScrolledText(master, wrap=tk.WORD, width=80, height=20)
        self.output.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        self.current_thread = None
        self.stdout_redirector = StdoutRedirector(self.output)
        sys.stdout = self.stdout_redirector
        self.output.insert(tk.END, cuda_message + "\n")
        self.entry.focus_set()

    def predict(self, event=None):
        selected_indices = self.listbox.curselection()
        selected_tickers = [self.sp500.iloc[i]['Ticker'] for i in selected_indices] if not self.sp500.empty else []
        manual_tickers = [s.strip().upper() for s in self.entry.get().split(',') if s.strip()]
        symbols = list(set(selected_tickers + manual_tickers))
        if not symbols:
            messagebox.showerror("Error", "Ingrese o seleccione al menos un ticker.")
            return
        if self.current_thread and self.current_thread.is_alive():
            messagebox.showinfo("Información", "Ya hay un análisis en progreso. Espere a que finalice.")
            return
        self.output.delete('1.0', tk.END)
        self.output.insert(tk.END, f"Iniciando predicción para: {', '.join(symbols)}...\n")
        self.progress['value'] = 0
        self.early_stopping_text.set("Early Stopping. Mejor: N/A | Actual: N/A | Peor: N/A")
        self.current_thread = threading.Thread(target=self.run_prediction, args=(symbols,), daemon=True)
        self.current_thread.start()

    def run_prediction(self, symbols):
        try:
            results = analyze_stocks(symbols, self.output, self.progress, self.early_stopping_text)
            if results:
                comp_file = generate_comparison_report(results)
                self.output.insert(tk.END, f"\nAnálisis completado. Informe comparativo guardado en: {comp_file}\n")
            else:
                self.output.insert(tk.END, "\nNo se completó el análisis para ningún ticker.\n")
            self.output.insert(tk.END, "\n¡Buen día, amigo!\n")
        except Exception as e:
            self.output.insert(tk.END, f"Error durante la predicción: {str(e)}\n")
            logging.error(f"Error en predicción: {str(e)}")
            logging.error(traceback.format_exc())

    def calibrate(self, event=None):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Error", "Seleccione un ticker para calibrar.")
            return
        symbol = self.sp500.iloc[selected_indices[0]]['Ticker']
        self.output.delete('1.0', tk.END)
        self.output.insert(tk.END, f"Iniciando calibración para {symbol}...\n")
        end_date = datetime.now()
        start_date = end_date - timedelta(days=3650)
        train_data, full_data, news = collect_data(symbol, start_date, end_date, API_KEY_FRED, self.output)
        if train_data is None or train_data.empty:
            self.output.insert(tk.END, f"No se pudieron obtener datos para {symbol}.\n")
            return
        try:
            X, y, scaler, sample_weights = preprocess_data(train_data, window_size=WINDOW_SIZE, forecast_horizon=FORECAST_HORIZON)
        except Exception as e:
            self.output.insert(tk.END, f"Error en preprocesamiento: {str(e)}\n")
            return
        param_grid = {'lstm_units': [150], 'dropout_rate': [0.3], 'learning_rate': [0.0005]}
        best_model, best_params, best_history = calibrate_model(X, y, sample_weights, param_grid, epochs=30, validation_split=0.1, batch_size=32, gui_output=self.output)
        self.output.insert(tk.END, f"Calibración completada para {symbol}. Mejores parámetros: {best_params}\n")

    def backtest(self, event=None):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Error", "Seleccione un ticker para backtesting.")
            return
        symbol = self.sp500.iloc[selected_indices[0]]['Ticker']
        self.output.delete('1.0', tk.END)
        self.output.insert(tk.END, f"Iniciando backtesting para {symbol}...\n")
        end_date = datetime.now()
        start_date = end_date - timedelta(days=3650)
        train_data, full_data, news = collect_data(symbol, start_date, end_date, API_KEY_FRED, self.output)
        if train_data is None or train_data.empty:
            self.output.insert(tk.END, f"No se pudieron obtener datos para {symbol}.\n")
            return
        try:
            X, y, scaler, sample_weights = preprocess_data(train_data, window_size=WINDOW_SIZE, forecast_horizon=FORECAST_HORIZON)
        except Exception as e:
            self.output.insert(tk.END, f"Error en preprocesamiento: {str(e)}\n")
            return
        model, X_test, y_test, history = train_model(X, y, sample_weights, self.output, self.progress, self.early_stopping_text)
        self.output.insert(tk.END, "Modelo entrenado para backtesting.\n")
        indices, predictions, actuals, rmse, mae, r2 = backtest_model(full_data, scaler, window_size=WINDOW_SIZE, model=model, step=1)
        self.output.insert(tk.END, f"Backtesting completado. RMSE: {rmse:.4f}, MAE: {mae:.4f}, R2: {r2:.4f}\n")
        backtest_fig = plot_backtest_results(indices, predictions, actuals, symbol)
        self.output.insert(tk.END, "Gráfico de backtesting generado.\n")

    def show_readme(self):
        readme_text = """
Predicción de Precios de Acciones

Este script utiliza deep learning para predecir precios futuros de acciones basándose en:
  - Datos históricos de precios, volumen y fundamentales (P/E y P/S).
  - Indicadores técnicos (SMA, EMA, RSI, MACD).
  - Datos macroeconómicos (tasa de interés, inflación, desempleo).
  - Análisis de sentimiento de noticias (FinBERT).

Nuevas mejoras:
  - Predicción multi-step para 1 día.
  - Uso de un conjunto de entrenamiento limpio (sin NaN) y conservación de la versión raw de News_Sentiment para graficar.
  - Cálculo de importancia mediante gradientes y sensibilidad por perturbación.
  - Gráficos en Excel, en particular "Sentimiento vs Precio" usando solo días con dato real.

Uso:
  1. Seleccione tickers o ingréselos manualmente (separados por comas).
  2. Presione "Predecir" para el análisis completo.
  3. Use "Calibrar" para ajustar hiperparámetros.
  4. Use "Backtest" para ejecutar un backtesting.
  5. Los resultados se guardan en la carpeta "Cotizaciones" en el escritorio.

Advertencia:
  Este script es experimental. Realice backtesting y validación antes de invertir.
        """
        messagebox.showinfo("Léame", readme_text)

def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    root = tk.Tk()
    gui = StockPredictionGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
